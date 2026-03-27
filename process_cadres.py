#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
四级正副干部评优名额确认工具

功能：
1. 按照汇总表第一个子表（结果汇总表）里的部门顺序
2. 去干部年度表里匹配对应部门的干部名单
3. 把这些干部明细填到汇总表的 2025年度 子表中
4. 自动统计每个部门的：考核基数、外派干部人数、评优名额
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import re


def get_department_mapping():
    """
    创建汇总表部门名称与干部年度表部门名称的映射关系
    """
    mapping = {
        '人力资源部': '人力资源部（含专职董事监事办公室、社保（年金）中心）',
        '企业发展部': '企业发展部（含运营监控中心）',
        '办公室': '办公室',
        '工会工作部': '工会工作部',
        '党建工作部': '党建工作部',
        '市场及客户服务部': '市场及客户服务部',
        '配网管理部': '配网管理部',
        '基建部': '基建部（含小基与迁改项目管理中心）',
        '资产管理部': '资产管理部',
        '财务部': '财务部',
        '创新与数字化部': '创新与数字化部',
        '新兴产业部': '新兴产业部',
        '法规部': '法规部',
        '监督部': '监督部',
        '十五运保电办': '十五运保电办',
        '公司党委巡察工作领导小组办公室': '党委巡察工作领导小组办公室（含巡察组）',
        '安全监管部': '安全监管部（含安全督查大队）',
        '审计部': '审计部',
        '龙岗供电局': '龙岗供电局',
        '宝安供电局': '宝安供电局',
        '福田供电局': '福田供电局',
        '坪山供电局': '坪山供电局',
        '罗湖供电局': '罗湖供电局',
        '龙华供电局': '龙华供电局',
        '南山供电局': '南山供电局',
        '光明供电局': '光明供电局',
        '大鹏供电局': '大鹏供电局',
        '深汕特别合作区供电局': '深汕供电局',
        '盐田供电局': '盐田供电局',
        '变电管理二所': '变电管理二所',
        '输电管理所': '输电管理所',
        '电力调度控制中心': '电力调度控制中心',
        '变电管理一所': '变电管理一所',
        '通信管理所': '通信管理所',
        '供应链服务中心': '供应链服务中心',
        '电力科学研究院': '电力科学研究院',
        '电网规划研究中心': '电网规划研究中心',
        '建设分公司': '建设分公司',
        '电力行政执法协助中心': '电力行政执法协助中心',
        '客户服务中心': '客户服务中心',
        '服务稽查中心': '服务稽查中心',
        '计量管理所': '计量管理所',
        '新闻中心': '新闻中心',
        '财务共享中心': '财务共享中心',
        '综合服务中心': '综合服务中心',
        '数字化与人工智能中心': '数字化与人工智能中心',
        '公司党校（人才发展中心）': '公司党校',
        '深圳市华睿欣能投资控股有限公司': '深圳市华睿欣能投资控股有限公司',
        '深圳前海蛇口自贸区供电有限公司': '深圳前海蛇口自贸区供电有限公司',
        '深圳电网智慧能源技术有限公司': '深圳电网智慧能源技术有限公司',
        '深圳市领康达服务有限公司': '深圳市领康达服务有限公司',
        '深圳南方电网深港科技创新有限公司': '深圳南方电网深港科技创新有限公司',
        '深圳低碳城供电有限公司': '深圳低碳城供电有限公司',
        '深圳市电力行业协会': '深圳市电力行业协会',
    }
    return mapping


def get_reverse_mapping():
    """
    创建反向映射：从干部年度表部门名称到汇总表部门名称
    """
    forward = get_department_mapping()
    reverse = {}
    for summary_name, annual_name in forward.items():
        reverse[annual_name] = summary_name
    # 处理直接匹配的部门
    direct_depts = [
        '办公室', '工会工作部', '党建工作部', '市场及客户服务部', '配网管理部',
        '资产管理部', '财务部', '创新与数字化部', '新兴产业部', '法规部',
        '监督部', '十五运保电办', '审计部', '龙岗供电局', '宝安供电局',
        '福田供电局', '坪山供电局', '罗湖供电局', '龙华供电局', '南山供电局',
        '光明供电局', '大鹏供电局', '盐田供电局', '变电管理二所', '输电管理所',
        '电力调度控制中心', '变电管理一所', '通信管理所', '供应链服务中心',
        '电力科学研究院', '电网规划研究中心', '建设分公司', '电力行政执法协助中心',
        '客户服务中心', '服务稽查中心', '计量管理所', '新闻中心', '财务共享中心',
        '综合服务中心', '数字化与人工智能中心', '深圳市电力行业协会',
        '深圳前海蛇口自贸区供电有限公司', '深圳电网智慧能源技术有限公司',
        '深圳市领康达服务有限公司', '深圳南方电网深港科技创新有限公司',
        '深圳低碳城供电有限公司'
    ]
    for dept in direct_depts:
        if dept not in reverse:
            reverse[dept] = dept
    return reverse


def calculate_excellent_quota(count, evaluation_result):
    """
    根据 rule.md 的规则计算评优名额
    
    参数:
        count: 干部人数
        evaluation_result: 领导班子评价结果 (优秀/良好/一般及以下)
    
    返回:
        评优名额
    """
    if count <= 0:
        return 0
    
    # 确定评价等级
    if evaluation_result in ['优秀', '良好']:
        level = 'good'  # 优秀或良好
    else:
        level = 'bad'  # 一般及以下
    
    # 根据人数和评价结果确定名额
    if count <= 2:
        if level == 'good':
            return 1
        else:
            return 0
    elif count == 3:
        if level == 'good':
            return 1
        else:
            return 0
    elif 4 <= count <= 6:
        if evaluation_result == '优秀':
            return 2
        elif evaluation_result == '良好':
            return 1
        else:
            return 0
    elif 7 <= count <= 10:
        if evaluation_result == '优秀':
            return 3
        elif evaluation_result == '良好':
            return 2
        else:
            return 1
    elif 11 <= count <= 13:
        if evaluation_result == '优秀':
            return 4
        elif evaluation_result == '良好':
            return 3
        else:
            return 1
    elif 14 <= count <= 16:
        if evaluation_result == '优秀':
            return 5
        elif evaluation_result == '良好':
            return 4
        else:
            return 2
    elif 17 <= count <= 19:
        if evaluation_result == '优秀':
            return 6
        elif evaluation_result == '良好':
            return 5
        else:
            return 2
    elif 20 <= count <= 22:
        if evaluation_result == '优秀':
            return 7
        elif evaluation_result == '良好':
            return 6
        else:
            return 3
    elif 23 <= count <= 25:
        if evaluation_result == '优秀':
            return 8
        elif evaluation_result == '良好':
            return 7
        else:
            return 3
    else:
        # 超过25人的情况，按规则外推
        if evaluation_result == '优秀':
            return min(count // 3, count)
        elif evaluation_result == '良好':
            return min(count // 3, count - 1)
        else:
            return min(count // 8, count)


def get_departments_from_summary(summary_path):
    """
    从汇总表第一个子表（结果汇总表）中提取部门列表及其评价结果
    保持原有顺序
    """
    df = pd.read_excel(summary_path, sheet_name='结果汇总表', header=None)
    
    departments = []
    eval_results = {}
    
    # 部门名称在列索引1，评价结果在列索引2
    # 从第4行开始（索引4），跳过分类标题行
    category_titles = ['综合型部门', '专业型部门', '支撑型部门', '监督型部门', 
                       '直属供电局', '其他直属单位——生产调度运维组',
                       '其他直属单位——生产服务支撑组', '其他直属单位——营销服务支撑组',
                       '参控股公司']
    
    for idx in range(4, len(df)):
        dept_name = df.iloc[idx, 1]
        eval_result = df.iloc[idx, 2]
        
        if pd.notna(dept_name) and dept_name not in category_titles:
            if isinstance(dept_name, str) and dept_name.strip():
                departments.append(dept_name.strip())
                if pd.notna(eval_result):
                    eval_results[dept_name.strip()] = str(eval_result).strip()
    
    return departments, eval_results


def read_annual_cadres(annual_path):
    """
    读取干部年度表中的干部信息
    """
    # 使用 openpyxl 读取原始数据，避免 pandas 的类型推断问题
    import openpyxl
    wb = openpyxl.load_workbook(annual_path, data_only=True)
    ws = wb['干部名册 ']
    
    cadres = []
    # 从第4行开始（索引4），跳过标题行
    for row in range(4, ws.max_row + 1):
        seq_no = ws.cell(row=row, column=1).value
        if seq_no is None or seq_no == '四级正干部（181人)':
            continue
        
        name = ws.cell(row=row, column=2).value
        position_code = ws.cell(row=row, column=3).value  # 职务代码
        rank = ws.cell(row=row, column=4).value  # 现职级
        dept_name = ws.cell(row=row, column=6).value  # 部门名称
        dept_order = ws.cell(row=row, column=7).value  # 部门职级排序
        external = ws.cell(row=row, column=8).value  # 外派标记
        
        if dept_name is not None:
            cadres.append({
                '序号': seq_no,
                '姓名': str(name) if name is not None else '',
                '职务': str(position_code) if position_code is not None else '',
                '现职级': rank,
                '部门名称': dept_name,
                '部门职级排序': dept_order,
                '外派': external
            })
    
    df = pd.DataFrame(cadres)
    return df


def match_department(summary_dept, reverse_mapping):
    """
    将汇总表部门名称匹配到干部年度表部门名称
    """
    if summary_dept in reverse_mapping:
        return reverse_mapping[summary_dept]
    
    # 尝试模糊匹配
    for annual_dept in reverse_mapping.keys():
        if summary_dept in annual_dept or annual_dept in summary_dept:
            return annual_dept
    
    return None


def process_and_fill_data(summary_path, annual_path, output_path):
    """
    主处理函数：填充2025年度子表并计算统计数据
    保留原始表格的合并单元格格式
    
    处理逻辑：
    1. 从2025年度子表获取原始的合并单元格结构和部门名称
    2. 将原始部门名称映射到汇总表部门名称
    3. 从汇总表获取部门评价结果
    4. 从干部年度表获取对应部门的干部名单
    5. 按照原始结构填充数据
    """
    # 获取部门映射
    forward_mapping = get_department_mapping()  # 汇总表 -> 干部年度表
    reverse_mapping = get_reverse_mapping()  # 干部年度表 -> 汇总表
    
    # 从汇总表获取部门评价结果
    _, eval_results = get_departments_from_summary(summary_path)
    
    # 读取干部年度数据
    df_annual = read_annual_cadres(annual_path)
    
    # 加载工作簿
    wb = load_workbook(summary_path)
    
    # 获取2025年度工作表
    if '2025年度' not in wb.sheetnames:
        print("错误：2025年度子表不存在")
        return {}
    
    ws_original = wb['2025年度']
    
    # 保存原始的合并单元格信息和部门名称
    original_structure = []  # [(start_row, end_row, original_dept_name)]
    for merged_range in sorted(ws_original.merged_cells.ranges, key=lambda x: x.min_row):
        if merged_range.min_col == 2 and merged_range.max_col == 2:
            start_row = merged_range.min_row
            end_row = merged_range.max_row
            # 获取原始的部门名称
            original_dept_name = ws_original.cell(row=start_row, column=2).value
            original_structure.append((start_row, end_row, original_dept_name))
    
    dept_stats = {}
    
    # 遍历原始结构，逐个填充
    for start_row, end_row, original_dept_name in original_structure:
        if original_dept_name is None:
            continue
        
        cadre_count = end_row - start_row + 1
        
        # 尝试将原始部门名称映射到汇总表部门名称
        # 首先尝试直接反向映射
        summary_dept = None
        
        # 检查是否在正向映射的值中（即干部年度表名称）
        for summ_dept, annual_dept in forward_mapping.items():
            if original_dept_name == annual_dept or original_dept_name.startswith(annual_dept):
                summary_dept = summ_dept
                break
        
        # 如果没找到，尝试模糊匹配
        if summary_dept is None:
            for summ_dept, annual_dept in forward_mapping.items():
                if summ_dept in original_dept_name or annual_dept in original_dept_name:
                    summary_dept = summ_dept
                    break
        
        # 如果还是没找到，尝试直接使用原名
        if summary_dept is None:
            # 检查是否直接就是汇总表部门名称
            if original_dept_name in forward_mapping or original_dept_name in reverse_mapping:
                summary_dept = original_dept_name
        
        if summary_dept is None:
            print(f"警告：无法映射部门 '{original_dept_name}'")
            continue
        
        # 获取对应的干部年度表部门名称
        annual_dept = forward_mapping.get(summary_dept, summary_dept)
        
        # 获取该部门的干部
        dept_cadres = df_annual[df_annual['部门名称'] == annual_dept].copy()
        
        # 如果没有找到，尝试使用原始部门名称作为年度表部门名称
        if len(dept_cadres) == 0:
            dept_cadres = df_annual[df_annual['部门名称'] == original_dept_name].copy()
        
        if len(dept_cadres) == 0:
            print(f"提示：部门 '{summary_dept}' ({original_dept_name}) 没有干部数据")
            # 清空该区域
            ws_original.cell(row=start_row, column=2, value=summary_dept)
            continue
        
        # 按部门职级排序
        dept_cadres = dept_cadres.sort_values('部门职级排序', key=lambda x: pd.to_numeric(x, errors='coerce'))
        
        # 只取与合并单元格行数相同的干部数量
        if len(dept_cadres) > cadre_count:
            dept_cadres = dept_cadres.head(cadre_count)
        
        # 统计
        total_count = len(dept_cadres)
        external_count = len(dept_cadres[dept_cadres['外派'].notna() & (dept_cadres['外派'] != '')])
        base_count = total_count - external_count  # 考核基数（不含外派干部）
        
        # 获取评价结果
        eval_result = eval_results.get(summary_dept, '/')
        
        # 计算评优名额
        excellent_quota = calculate_excellent_quota(base_count, eval_result)
        
        # 存储统计数据
        dept_stats[summary_dept] = {
            '考核基数': base_count,
            '外派干部人数': external_count,
            '评优名额': excellent_quota,
            '评价结果': eval_result
        }
        
        # 填充部门名称（只在第一个单元格填写）
        ws_original.cell(row=start_row, column=2, value=summary_dept)
        
        # 填充干部信息
        for i, (_, cadre) in enumerate(dept_cadres.iterrows()):
            row = start_row + i
            ws_original.cell(row=row, column=1, value=cadre['序号'])
            ws_original.cell(row=row, column=3, value=cadre['姓名'])
            ws_original.cell(row=row, column=4, value=cadre['职务'])
            ws_original.cell(row=row, column=5, value=cadre['现职级'])
            ws_original.cell(row=row, column=6, value=cadre['外派'] if cadre['外派'] else None)
    
    # 保存文件
    wb.save(output_path)
    print(f"已保存到：{output_path}")
    
    return dept_stats


def main():
    """
    主函数
    """
    import sys
    
    # 默认文件路径
    summary_path = '/workspace/公司党委管理四级正（副）干部综合考核评价情况汇总表.xlsx'
    annual_path = '/workspace/四级干部年度.xlsx'
    output_path = '/workspace/公司党委管理四级正（副）干部综合考核评价情况汇总表_已填充.xlsx'
    
    # 如果命令行提供了参数，使用命令行参数
    if len(sys.argv) >= 3:
        summary_path = sys.argv[1]
        annual_path = sys.argv[2]
    if len(sys.argv) >= 4:
        output_path = sys.argv[3]
    
    print("=" * 60)
    print("四级正副干部评优名额确认工具")
    print("=" * 60)
    print(f"\n汇总表路径：{summary_path}")
    print(f"干部年度表路径：{annual_path}")
    print(f"输出文件路径：{output_path}")
    print()
    
    # 处理数据
    dept_stats = process_and_fill_data(summary_path, annual_path, output_path)
    
    # 打印统计结果
    print("\n" + "=" * 60)
    print("各部门统计结果")
    print("=" * 60)
    print(f"{'部门名称':<30} {'考核基数':>8} {'外派人数':>8} {'评优名额':>8} {'评价结果':>10}")
    print("-" * 60)
    
    for dept_name, stats in dept_stats.items():
        print(f"{dept_name:<30} {stats['考核基数']:>8} {stats['外派干部人数']:>8} {stats['评优名额']:>8} {stats['评价结果']:>10}")
    
    print("=" * 60)
    print("处理完成！")


if __name__ == '__main__':
    main()
